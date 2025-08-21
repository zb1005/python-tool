from email.policy import default
import os
import pandas as pd


def read_excel_files(folder_path):
    # 获取文件夹中的所有文件
    all_files = os.listdir(folder_path)
    
    # 筛选出Excel文件，同时排除临时文件（以~$开头的文件）
    excel_files = [file for file in all_files if file.endswith(('.xlsx', '.xls','.XLSX','.XLS')) and not file.startswith('~$')]

    
    if not excel_files:
        print(f"在文件夹 {folder_path} 中未找到Excel文件")
        return None, None
    
    # 筛选出包含'物料冻结情况查询'的文件
    target_files = [file for file in excel_files if '物料冻结情况查询' in file]
    
    if not target_files:
        print(f"在文件夹 {folder_path} 中未找到命名包含'物料冻结情况查询'的Excel文件")
        return None, excel_files
    
    # 选择第一个匹配的文件作为基准df
    base_file = target_files[0]
    base_file_path = os.path.join(folder_path, base_file)
    
    try:
        base_df = pd.read_excel(base_file_path)
        print(f"已成功读取基准文件: {base_file}")
        # 从excel_files中移除基准文件，确保返回的excel_files不包含base_df对应的文件
        filtered_excel_files = [file for file in excel_files if file != base_file]
        return base_df, filtered_excel_files
        
    except Exception as e:
        print(f"读取文件 {base_file} 时出错: {e}")
        # 从excel_files中移除基准文件，确保返回的excel_files不包含base_df对应的文件
        filtered_excel_files = [file for file in excel_files if file != base_file]
        return None, filtered_excel_files

#计算每个物料的库存
def calculate_inventory(result_df,current_df):
    #将库存数量转换为浮点数
    current_df['非限制使用的库存'] = current_df['非限制使用的库存'].astype(float)
    current_df['Trans./Tfr'] = current_df['Trans./Tfr'].astype(float)
    inventory_mapping = {}
    for index, row in current_df.iterrows():
        inventory_mapping[row['物料编码']] = inventory_mapping.get(row['物料编码'],0) + row['非限制使用的库存'] + row['Trans./Tfr']
    result_df['库存'] = result_df['物料编码'].map(inventory_mapping)
    result_df['库存'] = result_df['库存'].fillna('无')
    return result_df


#对比base_df的物料编码列与当前df的物料编码列，将在base_df的物料编码列的行，列为是否有未结工单，取是，否则取否
def compare_material_codes_worklist(result_df, current_df):
    current_material_codes = set(current_df['物料编号'])
    result_df['是否有未结工单'] = result_df['物料编码'].apply(lambda x: '是' if x in current_material_codes else '否')
    return result_df

#对比base_df的物料编码列与当前df的物料编码列，将在base_df的物料编码列的行，列为是否有未结工单，取是，否则取否
def compare_material_codes_buylist(result_df, current_df):
    current_material_codes = set(current_df['物料编码'])
    result_df['是否有未清采购订单'] = result_df['物料编码'].apply(lambda x: '是' if x in current_material_codes else '否')
    return result_df

#对于当前的df建立，子项物料编码和是否在生效制造BOM中字段的映射，优先取是，且其父级未冻结、未停止生产；是，但其父级已冻结或停止生产；否
def MBOM_mapping(result_df,current_df):
    MBOM_mapping = {}
    for index, row in current_df.iterrows():
        if row['是否在生效制造BOM中'] == '是，且其父级未冻结、未停止生产' and MBOM_mapping.get(row['子项物料编码'],0) == 0:
            MBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
        if row['是否在生效制造BOM中'] == '是，但其父级已冻结或停止生产' and MBOM_mapping.get(row['子项物料编码'],0) == 0:
            MBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'
        if row['是否在生效制造BOM中'] == '否' and MBOM_mapping.get(row['子项物料编码'],0) == 0:
            MBOM_mapping[row['子项物料编码']] = '否'
    result_df['是否在生效MBOM中'] = result_df['物料编码'].map(MBOM_mapping)

    return result_df


#对于当前的df建立，子项物料编码和是否在生效制造BOM中字段的映射，优先取是，且其父级未冻结、未停止生产；是，但其父级已冻结或停止生产；否
def SBOM_mapping(result_df,current_df):
    SBOM_mapping = {}
    for index, row in current_df.iterrows():
        if row['是否在生效服务BOM中'] == '是，且其父级未冻结、未停止生产' and SBOM_mapping.get(row['子项物料编码'],0) == 0:
            SBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
            continue
        if row['是否在生效服务BOM中'] == '是，但其父级已冻结或停止生产' and SBOM_mapping.get(row['子项物料编码'],0) == 0:
            SBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'
            continue
        if row['是否在生效服务BOM中'] == '否' and SBOM_mapping.get(row['子项物料编码'],0) == 0:
            SBOM_mapping[row['子项物料编码']] = '否'
            continue
    result_df['是否在生效SBOM中'] = result_df['物料编码'].map(SBOM_mapping) 
    return result_df

#对于当前的df建立，子项物料编码和是否在生效制造BOM中字段的映射，优先取是，且其父级未冻结、未停止生产；是，但其父级已冻结或停止生产；否
def EBOM_mapping(result_df,current_df):
    EBOM_mapping = {}
    for index, row in current_df.iterrows():
        if row['备注'] == '是，且其父级未冻结、未停止生产' and EBOM_mapping.get(row['子项物料编码'],0) == 0:
            EBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
            continue
        if row['备注'] == '是，但其父级已冻结或停止生产' and EBOM_mapping.get(row['子项物料编码'],0) == 0:
            EBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'
            continue
        if row['备注'] == '否' and EBOM_mapping.get(row['子项物料编码'],0) == 0:
            EBOM_mapping[row['子项物料编码']] = '否'
            continue
    result_df['是否被EBOM引用'] = result_df['物料编码'].map(EBOM_mapping) 
    return result_df

#对于当前的df建立，子项物料编码和是否在生效制造BOM中字段的映射，优先取是，且其父级未冻结、未停止生产；是，但其父级已冻结或停止生产；否
def XBOM_mapping(result_df,current_df):
    XBOM_mapping = {}
    for index, row in current_df.iterrows():
        if row['是否在生效销售BOM中'] == '是，且其父级未冻结、未停止生产' and XBOM_mapping.get(row['子项物料编码'],0) == 0:
            XBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
            continue
        if row['是否在生效销售BOM中'] == '是，但其父级已冻结或停止生产' and XBOM_mapping.get(row['子项物料编码'],0) == 0:
            XBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'      
            continue
        if row['是否在生效销售BOM中'] == '否' and XBOM_mapping.get(row['子项物料编码'],0) == 0:
            XBOM_mapping[row['子项物料编码']] = '否'
            continue
    result_df['是否在生效XBOM中'] = result_df['物料编码'].map(XBOM_mapping) 
    return result_df

def last_trade_date(result_df,current_df):
    last_trade_date_mapping={}
    current_df['过帐日期'] = pd.to_datetime(current_df['过帐日期'],format='%Y-%m-%d')
    current_df['凭证日期'] = pd.to_datetime(current_df['凭证日期'],format='%Y-%m-%d')
    for index, row in current_df.iterrows():
        last_trade_date_mapping[row['物料编号']] = max(row['过帐日期'],row['凭证日期'])
    result_df['最后一次交易日期'] = result_df['物料编码'].map(last_trade_date_mapping).fillna('无')
    return result_df

def service_product(result_df,current_df):
    service_product_mapping={}
    replace_product_mapping={}
    replace_relation_mapping={}
    current_df['替换物料'] = current_df['替换物料'].map(lambda x: '无' if x == 'nan' else str(int(float(x))))
    current_df['替换关系描述'] = current_df['替换关系描述'].map(lambda x: '无' if x == 'nan' else x)

    for index, row in current_df.iterrows():
        service_product_mapping[row['物料编码']] = row['是否服务物料']
        if '替换物料' in current_df.columns:
            replace_product_mapping[row['物料编码']] = row['替换物料']
        if '替换关系描述' in current_df.columns:
            replace_relation_mapping[row['物料编码']] = row['替换关系描述']
    result_df['是否服务物料'] = result_df['物料编码'].map(service_product_mapping)

    if '替换物料' in current_df.columns:
        result_df['替换物料'] = result_df['物料编码'].map(replace_product_mapping).fillna('无')
        result_df['替换物料'] = result_df['替换物料'].fillna('无')
    if '替换关系描述' in current_df.columns:
        result_df['替换关系描述'] = result_df['物料编码'].map(replace_relation_mapping).fillna('无')
        result_df['替换关系描述'] = result_df['替换关系描述'].fillna('无')

    return result_df

# 示例用法
if __name__ == "__main__":
    # 替换为实际的文件夹路径
    folder_path = r"C:\Users\zhangbon\Desktop\2025.08.06--物料冻结条件查询--杨亨亨"
    base_df, all_excel_files = read_excel_files(folder_path)
    result_df = base_df[["物料编码","物料描述"]].copy()
    result_df = result_df.astype(str)
    
    print(all_excel_files)

    for excel_file in all_excel_files:
        current_path = os.path.join(folder_path,excel_file)
        current_df = pd.read_excel(current_path)
        current_df = current_df.astype(str)

        if len(current_df.columns) >= 7:
            if (current_df.columns[:7]==['物料编码','物料描述','工厂','存储位置','仓储地点的描述','非限制使用的库存','Trans./Tfr']).all():
                result_df = calculate_inventory(result_df, current_df)
                continue

        if (current_df.columns[:4] == ['工厂','订单类型','订单','物料编号']).all():
            result_df = compare_material_codes_worklist(result_df, current_df)
            continue
        if (current_df.columns[:4] == ['凭证日期','采购凭证','物料编码','物料描述']).all():
            result_df = compare_material_codes_buylist(result_df, current_df)
            continue
        if '是否在生效制造BOM中' in current_df.columns:
            result_df = MBOM_mapping(result_df, current_df)
            continue
        if '是否在生效服务BOM中' in current_df.columns:
            result_df = SBOM_mapping(result_df, current_df)
            continue
        if len(current_df.columns) >= 7:
            if (current_df.columns[:7]==['子项物料编码','子项物料描述','父项物料编码','父项物料描述','父物料是否冻结','父物料的产品生命周期状态','备注']).all():
                result_df = EBOM_mapping(result_df, current_df)
                continue
        if '是否在生效服务BOM中' in current_df.columns:
            result_df = XBOM_mapping(result_df, current_df)
            continue
        if (current_df.columns[:4] == ['物料编号','物料描述','过帐日期','凭证日期']).all():
            result_df = last_trade_date(result_df, current_df)
            continue
        if '是否服务物料' in current_df.columns:
            result_df = service_product(result_df, current_df)
            continue
    #规整reslut_df的列名顺序为['物料编码','物料描述','库存','是否有未结工单','是否有未清采购订单','是否在生效MBOM中','是否在生效SBOM中','是否在生效XBOM引用','是否被EBOM引用','最后一次交易日期','是否服务物料','替换物料','替换关系描述']，如果有缺少的列名则自动跳过
    columns = ['物料编码','物料描述','库存','是否有未结工单','是否有未清采购订单','是否在生效MBOM中','是否在生效SBOM中','是否在生效XBOM中','是否被EBOM引用','最后一次交易日期','是否服务物料','替换物料','替换关系描述']

    for column in columns:
        if column not in result_df.columns:
            result_df[column] = '无'
    result_df = result_df[columns]
    result_df['是否可冻结'] = '/'
    for index,row in result_df.iterrows():
        #如果库存为无，没有未结工单，没有未清采购订单，且不是服务物料，不在生效BOM中或者是，但其父级已冻结或停止生产，并且不是服务物料，则该物料在是否可冻结列的值取是
        if row['库存'] == '无' and row['是否有未结工单'] == '否' and row['是否有未清采购订单'] == '否' and row['是否服务物料'] == '否' and (row['是否在生效MBOM中'] in ['否','是否在生效MBOM中']) and (row['是否在生效SBOM中'] in ['否','是否在生效SBOM中']) and (row['是否在生效XBOM中'] in ['无','否','是否在生效SBOM中']) and (row['是否被EBOM引用'] in ['否','是，但其父级已冻结或停止生产']):
            result_df.loc[index,'是否可冻结'] = '是'

        else:
            result_df.loc[index,'是否可冻结'] = '否'
    
    # 创建带颜色标记的Excel文件
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # 定义颜色填充
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    # 创建新的工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物料冻结分析"
    
    # 将DataFrame写入工作表
    for r in dataframe_to_rows(result_df, index=False, header=True):
        ws.append(r)
    
    # 获取列名对应的列索引
    col_indices = {cell.value: cell.column for cell in ws[1]}
    
    # 应用颜色规则
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
        excel_row = row
        data_row = result_df.iloc[row_idx-2]
        
        # 规则1: 如果是否可冻结字段的值为否，则标红当前行的物料编码
        if data_row['是否可冻结'] == '否':
            if '物料编码' in col_indices:
                excel_row[col_indices['物料编码']-1].fill = red_fill
        
        # 规则2: 如果库存不为无，则标红该行的库存
        if data_row['库存'] != '无':
            if '库存' in col_indices:
                excel_row[col_indices['库存']-1].fill = red_fill
        
        # 规则3: 如果是否有未结工单不为否，则标红
        if data_row['是否有未结工单'] != '否':
            if '是否有未结工单' in col_indices:
                excel_row[col_indices['是否有未结工单']-1].fill = red_fill
        
        # 规则4: 如果是否有未清采购订单不为否，则标红
        if data_row['是否有未清采购订单'] != '否':
            if '是否有未清采购订单' in col_indices:
                excel_row[col_indices['是否有未清采购订单']-1].fill = red_fill
        
        # 规则5: 如果是否在生效MBOM中的值为是，且其父级未冻结、未停止生产，则标红
        mbom_value = str(data_row['是否在生效MBOM中'])
        if '是，且其父级未冻结、未停止生产' in mbom_value:
            if '是否在生效MBOM中' in col_indices:
                excel_row[col_indices['是否在生效MBOM中']-1].fill = red_fill
        
        # 规则6: 如果是否在生效SBOM中的值为是，且其父级未冻结、未停止生产，则标红
        sbom_value = str(data_row['是否在生效SBOM中'])
        if '是，且其父级未冻结、未停止生产' in sbom_value:
            if '是否在生效SBOM中' in col_indices:
                excel_row[col_indices['是否在生效SBOM中']-1].fill = red_fill
        
        # 规则7: 如果是否在生效XBOM中的值为是，且其父级未冻结、未停止生产，则标红
        xbom_value = str(data_row['是否在生效XBOM中'])
        if '是，且其父级未冻结、未停止生产' in xbom_value:
            if '是否在生效XBOM中' in col_indices:
                excel_row[col_indices['是否在生效XBOM中']-1].fill = red_fill
        
        # 规则8: 如果是否被EBOM引用的值为是，且其父级未冻结、未停止生产，则标红
        ebom_value = str(data_row['是否被EBOM引用'])
        if '是，且其父级未冻结、未停止生产' in ebom_value:
            if '是否被EBOM引用' in col_indices:
                excel_row[col_indices['是否被EBOM引用']-1].fill = red_fill
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存带颜色的Excel文件
    output_path = r'C:\Users\zhangbon\Desktop\物料-冻结-数据处理结果-带颜色.xlsx'
    wb.save(output_path)
    print(f"带颜色标记的结果已保存到: {output_path}")
    
    # 同时保存普通版本
    result_df.to_excel(r'C:\Users\zhangbon\Desktop\物料-冻结-数据处理结果.xlsx', index=False)



