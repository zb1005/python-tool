"""
Excel数据核对工具 (Part 1/3)
功能：一对多数据核对、自动处理列名格式、主键拼接比对
环境要求：Python 3.7+，需安装 pandas openpyxl
安装依赖：pip install pandas openpyxl
"""

import pandas as pd
import os
from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl import load_workbook

# ---------------------- 工具函数 ----------------------
def clean_column_names(df):
    """清洗列名：去除前后空格和换行符"""
    df.columns = df.columns.str.strip().str.replace('\n', '')
    return df

def generate_composite_key(row, key_columns):
    """生成组合主键（处理空值和类型转换）"""
    return '_'.join([
        str(row[col]).strip() if pd.notna(row[col]) and str(row[col]).strip() != ''
        else '<空值>'  # 明确标记空值
        for col in key_columns
    ])

def read_excel_safely(file_path, key_columns, sheet_name=0):
    """
    安全读取Excel文件
    :return: 处理后的DataFrame
    """
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            dtype=str,
            keep_default_na=False,
            engine='openpyxl'
        )
        df = clean_column_names(df)
        missing_keys = [col for col in key_columns if col not in df.columns]
        if missing_keys:
            raise ValueError(f"缺失主键列: {', '.join(missing_keys)}")
        df['_composite_key'] = df.apply(
            lambda x: generate_composite_key(x, key_columns),
            axis=1
        )
        return df
    except Exception as e:
        raise ValueError(f"文件读取失败 [{os.path.basename(file_path)}]: {str(e)}")


"""
Excel数据核对工具 (Part 2/3)
"""

from datetime import datetime
import re


def normalize_date(date_str):
    """日期格式标准化（支持多种格式解析）"""
    if pd.isna(date_str) or str(date_str).strip() in ['', '<空值>']:
        return None

    formats = [
        '%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y',
        '%Y%m%d', '%Y.%m.%d', '%d-%b-%y'
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(str(date_str).strip(), fmt)
            return dt.date().isoformat()  # 统一转为YYYY-MM-DD格式
        except ValueError:
            continue
    return str(date_str).strip()  # 无法解析时返回原始值


def normalize_phone(phone_str):
    """电话号码归一化处理"""
    if pd.isna(phone_str) or str(phone_str).strip() in ['', '<空值>']:
        return None

    # 去除所有非数字字符
    cleaned = re.sub(r'\D', '', str(phone_str))

    # 处理国际区号
    if cleaned.startswith('86') and len(cleaned) > 11:
        return cleaned[2:]  # 去除86前缀
    elif cleaned.startswith('0086'):
        return cleaned[4:]
    elif len(cleaned) == 11 and cleaned.startswith('1'):
        return cleaned
    return cleaned

def normalize_number(num_str):
    """数值格式归一化处理（将4.00、4.000和4视为相同值）"""
    if pd.isna(num_str) or str(num_str).strip() in ['', '<空值>']:
        return None
    
    try:
        # 尝试转换为浮点数再转回字符串，去除多余的0和小数点
        num = float(str(num_str).strip())
        if num.is_integer():
            return str(int(num))
        return str(num)
    except ValueError:
        # 如果无法转换为数字，返回原始值
        return str(num_str).strip()

def normalize_string(str_value):
    """字符串归一化处理（去除两侧空格）"""
    if pd.isna(str_value) or str(str_value).strip() in ['', '<空值>']:
        return ""
    return str(str_value).strip()

def compare_datasets(source_df, target_df, key_columns):
    """处理标准格式下的目标数据"""
    if target_df.iloc[0, 0] == '元数据标准名称':
        print('==========标准模板处理===========')
        target_df.columns = target_df.iloc[0].tolist()
        target_df = target_df.drop([0, 1]).reset_index(drop=True)
        target_df = target_df.drop(target_df.columns[0], axis=1)
        target_df.reset_index(drop=True, inplace=True)

    """精准核对函数（仅核对共同字段）"""
    # 生成主键列
    source_df['_composite_key'] = source_df.apply(lambda x: generate_composite_key(x, key_columns), axis=1)
    target_df['_composite_key'] = target_df.apply(lambda x: generate_composite_key(x, key_columns), axis=1)

    # 获取主键集合
    source_keys = set(source_df['_composite_key'])
    target_keys = set(target_df['_composite_key'])
    common_keys = source_keys & target_keys

    # 获取共同字段（排除主键列）
    common_columns = list(set(source_df.columns) & set(target_df.columns) - {'_composite_key'})

    results = []

    # 处理仅存在于源数据的主键
    for key in source_keys - target_keys:
        results.append({
            '主键状态': '仅存在于源文件',
            '组合主键': key,
            **{f"源_{col}": source_df[source_df['_composite_key'] == key].iloc[0][col] for col in common_columns},
            **{f"目标_{col}": "" for col in common_columns}
        })

    # 处理仅存在于目标数据的主键
    for key in target_keys - source_keys:
        results.append({
            '主键状态': '仅存在于目标文件',
            '组合主键': key,
            **{f"源_{col}": "" for col in common_columns},
            **{f"目标_{col}": target_df[target_df['_composite_key'] == key].iloc[0][col] for col in common_columns}
        })

    # 处理共同主键的数据差异
    for key in common_keys:
        src_row = source_df[source_df['_composite_key'] == key].iloc[0]
        tgt_row = target_df[target_df['_composite_key'] == key].iloc[0]

        diff_details = {}
        row_data = {
            '主键状态': '数据一致',
            '组合主键': key
        }

        # 仅比对共同字段
        for col in common_columns:
            src_val = src_row[col]
            tgt_val = tgt_row[col]

            # 特殊字段预处理
            normalized_src = src_val
            normalized_tgt = tgt_val

            # 日期字段处理（列名包含date/日期）
            if 'date' in col.lower() or '日期' in col:
                normalized_src = normalize_date(src_val)
                normalized_tgt = normalize_date(tgt_val)

            # 电话字段处理（列名包含phone/电话）
            elif 'phone' in col.lower() or '电话' in col or '手机号' in col:
                normalized_src = normalize_phone(src_val)
                normalized_tgt = normalize_phone(tgt_val)

            # 数值字段处理（列名包含number/数值/数量等）
           
            normalized_src = normalize_number(normalized_src)
            normalized_tgt = normalize_number(normalized_tgt)
                
            # 对所有数据去除两侧空格
            normalized_src = normalize_string(normalized_src)
            normalized_tgt = normalize_string(normalized_tgt)

            # 空值等价处理
            src_empty = pd.isna(normalized_src) or str(normalized_src).strip() in ['', '<空值>']
            tgt_empty = pd.isna(normalized_tgt) or str(normalized_tgt).strip() in ['', '<空值>']

            # 实际比对逻辑
            if src_empty and tgt_empty:
                status = '一致'
            elif normalized_src == normalized_tgt:
                status = '一致'
            else:
                status = '不一致'
                diff_details[col] = {
                    '源值': src_val,
                    '目标值': tgt_val
                }

            # 记录原始值
            row_data[f"源_{col}"] = src_val
            row_data[f"目标_{col}"] = tgt_val

        # 更新状态
        if diff_details:
            row_data['主键状态'] = f"发现{len(diff_details)}处差异"
            row_data['差异详情'] = str(diff_details)
            row_data['差异列名'] = ",".join(diff_details.keys())
        else:
            row_data['差异详情'] = None

        results.append(row_data)

    return pd.DataFrame(results)


"""
Excel数据核对工具 (Part 3/3)
"""

def generate_detailed_report(result_df, output_path):
    """生成带精准颜色标注的报告"""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 写入数据
        result_df.to_excel(writer, index=False, sheet_name='核对结果')

        workbook = writer.book
        worksheet = writer.sheets['核对结果']

        # 定义颜色
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 亮黄色
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 纯红色

        # 获取关键列位置
        status_col_idx = 0  # 主键状态列位置（第1列）
        diff_col_idx = len(result_df.columns) - 1  # 差异详情列位置（最后1列）

        # 遍历所有行（从第2行开始）
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            status_cell = row[status_col_idx]
            status_value = status_cell.value

            # 仅存在于源/目标文件的整行标黄
            if status_value in ['仅存在于源文件', '仅存在于目标文件']:
                for cell in row:
                    cell.fill = yellow_fill

            # 存在差异的行标红关键列
            elif '差异' in str(status_value):
                # 标红主键状态列
                status_cell.fill = red_fill

                # 标红差异详情列
                diff_cell = row[diff_col_idx]
                diff_cell.fill = red_fill

                # 获取差异详情
                diff_details = result_df.iloc[row_idx - 2]['差异详情']
                if diff_details and isinstance(diff_details, str):
                    # 解析差异详情
                    import ast
                    diff_details = ast.literal_eval(diff_details)
                    all_keys = list(diff_details.keys())
                    # 遍历所有列，标红存在差异的字段
                    for col_idx, cell in enumerate(row):
                        col_name = result_df.columns[col_idx]
                        for key in all_keys:
                            if key in col_name:
                                cell.fill = red_fill

        # 设置自适应列宽
        for col in worksheet.columns:
            max_length = 5
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length)
            worksheet.column_dimensions[column].width = adjusted_width

        # 冻结首行
        worksheet.freeze_panes = 'A2'


def batch_compare(source_path, target_paths, key_columns, output_dir):
    """批量比对主程序"""
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    try:
        # 读取源数据
        print(f"\n{'=' * 30} 开始处理 {'=' * 30}")
        print(f"[{datetime.now().strftime('%H:%M:%S')}] 正在加载源文件: {os.path.basename(source_path)}")
        source_df = read_excel_safely(source_path, key_columns)
        print(f"源文件加载成功，记录数: {len(source_df)}")
        print(f"源数据列名为{source_df.columns}")

        # 遍历目标文件
        for target_path in target_paths:
            start_time = datetime.now()
            target_name = os.path.basename(target_path)
            print(f"\n[{start_time.strftime('%H:%M:%S')}] 正在处理目标文件: {target_name}")

            try:
                # 读取目标数据
                target_df = read_excel_safely(target_path, key_columns)
                print(f"目标文件加载成功，记录数: {len(target_df)}")
                print(f"目标数据列名为{target_df.columns}")

                # 执行比对
                result_df = compare_datasets(source_df, target_df, key_columns)

                # 生成报告文件名
                report_name = f"比对报告_{os.path.splitext(os.path.basename(source_path))[0]}_vs_{os.path.splitext(target_name)[0]}.xlsx"
                output_path = os.path.join(output_dir, report_name)

                # 生成报告
                generate_detailed_report(result_df, output_path)

                # 打印统计信息
                duration = (datetime.now() - start_time).total_seconds()
                print(f"[√] 完成比对 ({duration:.2f}秒)")
                print(f"差异统计：")
                print(f"  仅存在于源文件: {len(result_df[result_df['主键状态'] == '仅存在于源文件'])}")
                print(f"  仅存在于目标文件: {len(result_df[result_df['主键状态'] == '仅存在于目标文件'])}")
                print(f"  存在数据差异: {len(result_df[result_df['主键状态'].str.contains('差异')])}")
                print(f"报告已保存至: {output_path}")

            except Exception as e:
                print(f"[×] 处理失败: {str(e)}")
                continue

    except Exception as e:
        print(f"\n[!] 严重错误: {str(e)}")
    finally:
        print(f"\n{'=' * 30} 处理完成 {'=' * 30}")


if __name__ == "__main__":
    # ===================== 配置区 =====================
    # 使用原始字符串处理Windows路径
    SOURCE_FILE = r"C:\Users\zhangbon\Desktop\案例\核对数据集\MDM中销售BOM.XLSX"
    TARGET_FILES = [
r"C:\Users\zhangbon\Desktop\案例\核对数据集\ERP销售BOM.XLSX"
    ]
    KEY_COLUMNS = ["工厂","物料编码","组件"]  # 主键列
    OUTPUT_DIR = r"C:\Users\zhangbon\Desktop\案例\核对数据集\输出报告"
    # ================================================

    # 执行批量比对
    batch_compare(
        source_path=SOURCE_FILE,
        target_paths=TARGET_FILES,
        key_columns=KEY_COLUMNS,
        output_dir=OUTPUT_DIR
    )
