import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext, messagebox
import threading
from datetime import datetime

class MaterialFreezeGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("物料冻结分析工具")
        self.root.geometry("1200x800")
        
        # 设置中文字体
        self.root.option_add("*Font", "SimSun 10")
        
        # 初始化变量
        self.folder_path = tk.StringVar()
        self.result_df = None
        self.processing = False
        
        self.setup_ui()
        
    def setup_ui(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="物料冻结分析工具", 
                               font=("SimHei", 16, "bold"))
        title_label.pack(pady=(0, 20))
        
        # 文件夹选择区域
        folder_frame = ttk.LabelFrame(main_frame, text="数据文件夹选择", padding="10")
        folder_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(folder_frame, text="文件夹路径:").pack(side=tk.LEFT, padx=(0, 10))
        folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_path, width=60)
        folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        ttk.Button(folder_frame, text="浏览...", command=self.browse_folder).pack(side=tk.LEFT)
        
        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.start_button = ttk.Button(button_frame, text="开始处理", command=self.start_processing)
        self.start_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.export_button = ttk.Button(button_frame, text="导出结果", command=self.export_results, state=tk.DISABLED)
        self.export_button.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(button_frame, text="清空日志", command=self.clear_log).pack(side=tk.LEFT)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, mode='determinate')
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        
        # 分割窗口
        paned_window = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True)
        
        # 日志区域
        log_frame = ttk.LabelFrame(paned_window, text="处理日志", padding="5")
        paned_window.add(log_frame, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=50)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 结果预览区域
        result_frame = ttk.LabelFrame(paned_window, text="结果预览", padding="5")
        paned_window.add(result_frame, weight=2)
        
        # 创建表格
        self.setup_result_table(result_frame)
        
    def setup_result_table(self, parent):
        # 创建Treeview
        columns = ['物料编码', '物料描述', '库存', '是否有未结工单', '是否有未清采购订单', 
                  '是否在生效MBOM中', '是否在生效SBOM中', '是否在生效XBOM中', 
                  '是否被EBOM引用', '是否可冻结']
        
        self.tree = ttk.Treeview(parent, columns=columns, show='headings', height=15)
        
        # 设置列标题和宽度
        column_widths = {
            '物料编码': 100, '物料描述': 150, '库存': 80, '是否有未结工单': 100,
            '是否有未清采购订单': 100, '是否在生效MBOM中': 120, '是否在生效SBOM中': 120,
            '是否在生效XBOM中': 120, '是否被EBOM引用': 120, '是否可冻结': 80
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths.get(col, 100), anchor='center')
        
        # 添加滚动条
        scrollbar_y = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=self.tree.xview)
        
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # 布局
        self.tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_y.grid(row=0, column=1, sticky='ns')
        scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        
    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path.set(folder_selected)
            self.log_message(f"已选择文件夹: {folder_selected}")
            
    def log_message(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def clear_log(self):
        self.log_text.delete(1.0, tk.END)
        
    def start_processing(self):
        folder_path = self.folder_path.get()
        if not folder_path:
            messagebox.showwarning("警告", "请先选择数据文件夹")
            return
            
        if not os.path.exists(folder_path):
            messagebox.showerror("错误", "文件夹不存在")
            return
            
        if self.processing:
            messagebox.showinfo("提示", "处理已在进行中")
            return
            
        # 禁用按钮并开始处理
        self.start_button.config(state=tk.DISABLED)
        self.export_button.config(state=tk.DISABLED)
        self.processing = True
        
        # 清空之前的表格数据
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # 在新线程中处理数据
        thread = threading.Thread(target=self.process_data, args=(folder_path,))
        thread.daemon = True
        thread.start()
        
    def process_data(self, folder_path):
        try:
            self.log_message("开始处理数据...")
            
            # 读取Excel文件
            base_df, all_excel_files = self.read_excel_files(folder_path)
            if base_df is None:
                self.log_message("未找到有效的基准文件")
                return
                
            self.log_message(f"找到 {len(all_excel_files)} 个辅助Excel文件")
            
            result_df = base_df[["物料编码", "物料描述"]].copy()
            result_df = result_df.astype(str)
            result_df['库存'] = 0
            result_df['是否有未结工单'] = '未找到'
            result_df['是否有未清采购订单'] = '未找到'
            result_df['最后一次交易日期'] = '未找到'
            
            total_files = len(all_excel_files)
            
            for i, excel_file in enumerate(all_excel_files):
                current_path = os.path.join(folder_path, excel_file)
                self.log_message(f"处理文件: {excel_file}")
                
                try:
                    current_df = pd.read_excel(current_path)
                    current_df = current_df.astype(str)
                    
                    # 根据列名判断文件类型并处理
                    if len(current_df.columns) >= 7:
                        if (current_df.columns[:7] == ['物料编码', '物料描述', '工厂', '存储位置', '仓储地点的描述', '非限制使用的库存', 'Trans./Tfr']).all():
                            result_df = self.calculate_inventory(result_df, current_df)
                            continue
                            
                    if (current_df.columns[:4] == ['工厂', '订单类型', '订单', '物料编号']).all():
                        result_df = self.compare_material_codes_worklist(result_df, current_df)
                        continue
                        
                    if (current_df.columns[:4] == ['凭证日期', '采购凭证', '物料编码', '物料描述']).all():
                        result_df = self.compare_material_codes_buylist(result_df, current_df)
                        continue
                        
                    if '是否在生效制造BOM中' in current_df.columns:
                        result_df = self.MBOM_mapping(result_df, current_df)
                        continue
                        
                    if '是否在生效服务BOM中' in current_df.columns:
                        result_df = self.SBOM_mapping(result_df, current_df)
                        continue
                        
                    if len(current_df.columns) >= 7:
                        if (current_df.columns[:7] == ['子项物料编码', '子项物料描述', '父项物料编码', '父项物料描述', '父物料是否冻结', '父物料的产品生命周期状态', '备注']).all():
                            result_df = self.EBOM_mapping(result_df, current_df)
                            continue
                            
                    if '是否在生效销售BOM中' in current_df.columns:
                        result_df = self.XBOM_mapping(result_df, current_df)
                        continue
                        
                    if (current_df.columns[:4] == ['物料编号', '物料描述', '过帐日期', '凭证日期']).all():
                        result_df = self.last_trade_date(result_df, current_df)
                        continue
                        
                    if '是否服务物料' in current_df.columns:
                        result_df = self.service_product(result_df, current_df)
                        continue
                        
                    self.log_message(f"无法识别文件格式: {excel_file}")
                    
                except Exception as e:
                    self.log_message(f"处理文件 {excel_file} 时出错: {str(e)}")
                    
                # 更新进度
                progress = ((i + 1) / total_files) * 100
                self.root.after(0, lambda p=progress: self.progress_var.set(p))
                
            # 规整列名顺序
            columns = ['物料编码', '物料描述', '库存', '是否有未结工单', '是否有未清采购订单', 
                      '是否在生效MBOM中', '是否在生效SBOM中', '是否在生效XBOM中', 
                      '是否被EBOM引用', '最后一次交易日期', '是否服务物料', '替换物料', '替换关系描述']
            
            for column in columns:
                if column not in result_df.columns:
                    result_df[column] = '无'
            result_df = result_df[columns]
            result_df['是否可冻结'] = '/'
            result_df['库存'] = result_df['库存'].apply(lambda x : '无' if x == 0 else x).astype(str)
            result_df['是否有未结工单'] = result_df['是否有未结工单'].apply(lambda x : '否' if x == '未找到' else x)
            result_df['是否有未清采购订单'] = result_df['是否有未清采购订单'].apply(lambda x : '否' if x == '未找到' else x)
            result_df['最后一次交易日期'] = result_df['最后一次交易日期'].apply(lambda x : '无交易' if x == '未找到' else x)
            # 计算是否可冻结
            for index, row in result_df.iterrows():
                if row['库存'] == '无' and row['是否有未结工单'] == '否' and row['是否有未清采购订单'] == '否' and (row['是否在生效MBOM中'] in ['否', '是，但其父级已冻结或停止生产']) and (row['是否在生效SBOM中'] in ['否', '是，但其父级已冻结或停止生产']) and (row['是否在生效XBOM中'] in ['否', '是，但其父级已冻结或停止生产']) and (row['是否被EBOM引用'] in ['否', '是，但其父级已冻结或停止生产']):
                    result_df.loc[index, '是否可冻结'] = '是'
                else:
                    result_df.loc[index, '是否可冻结'] = '否'
                    
            self.result_df = result_df
            
            # 显示结果
            self.root.after(0, self.display_results)
            self.log_message("数据处理完成")
            
        except Exception as e:
            self.log_message(f"处理过程中出错: {str(e)}")
            messagebox.showerror("错误", f"处理过程中出错: {str(e)}")
            
        finally:
            self.processing = False
            self.root.after(0, lambda: self.start_button.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.export_button.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.progress_var.set(0))
            
    def display_results(self):
        if self.result_df is None:
            return
            
        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # 显示主要列
        display_columns = ['物料编码', '物料描述', '库存', '是否有未结工单', '是否有未清采购订单', 
                          '是否在生效MBOM中', '是否在生效SBOM中', '是否在生效XBOM中', 
                          '是否被EBOM引用', '是否可冻结']
        
        # 插入数据
        for _, row in self.result_df.iterrows():
            values = [str(row[col]) for col in display_columns]
            self.tree.insert('', tk.END, values=values)
            
        # 统计信息
        total_count = len(self.result_df)
        freezable_count = len(self.result_df[self.result_df['是否可冻结'] == '是'])
        
        self.log_message(f"总计处理物料: {total_count} 个")
        self.log_message(f"可冻结物料: {freezable_count} 个")
        self.log_message(f"不可冻结物料: {total_count - freezable_count} 个")
        
    def export_results(self):
        if self.result_df is None:
            messagebox.showwarning("警告", "没有可导出的结果")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            title='保存结果'
        )
        
        if file_path:
            try:
                # 创建带颜色标记的Excel文件
                import openpyxl
                from openpyxl.styles import PatternFill
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # 定义颜色填充
                red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                
                # 创建新的工作簿和工作表
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "物料冻结分析"
                
                # 将DataFrame写入工作表
                for r in dataframe_to_rows(self.result_df, index=False, header=True):
                    ws.append(r)
                
                # 获取列名对应的列索引
                col_indices = {cell.value: cell.column for cell in ws[1]}
                
                # 应用颜色规则
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
                    excel_row = row
                    data_row = self.result_df.iloc[row_idx-2]
                    
                    # 规则1: 如果是否可冻结字段的值为否，则标红当前行的物料编码
                    if str(data_row.get('是否可冻结', '')) == '否':
                        if '物料编码' in col_indices:
                            excel_row[col_indices['物料编码']-1].fill = red_fill
                    
                    # 规则2: 如果库存不为无，则标红该行的库存
                    if str(data_row.get('库存', '')) != '无':
                        if '库存' in col_indices:
                            excel_row[col_indices['库存']-1].fill = red_fill
                    
                    # 规则3: 如果是否有未结工单不为否，则标红
                    if str(data_row.get('是否有未结工单', '')) != '否':
                        if '是否有未结工单' in col_indices:
                            excel_row[col_indices['是否有未结工单']-1].fill = red_fill
                    
                    # 规则4: 如果是否有未清采购订单不为否，则标红
                    if str(data_row.get('是否有未清采购订单', '')) != '否':
                        if '是否有未清采购订单' in col_indices:
                            excel_row[col_indices['是否有未清采购订单']-1].fill = red_fill
                    
                    # 规则5: 如果是否在生效MBOM中的值为是，且其父级未冻结、未停止生产，则标红
                    mbom_value = str(data_row.get('是否在生效MBOM中', ''))
                    if '是，且其父级未冻结、未停止生产' in mbom_value:
                        if '是否在生效MBOM中' in col_indices:
                            excel_row[col_indices['是否在生效MBOM中']-1].fill = red_fill
                    
                    # 规则6: 如果是否在生效SBOM中的值为是，且其父级未冻结、未停止生产，则标红
                    sbom_value = str(data_row.get('是否在生效SBOM中', ''))
                    if '是，且其父级未冻结、未停止生产' in sbom_value:
                        if '是否在生效SBOM中' in col_indices:
                            excel_row[col_indices['是否在生效SBOM中']-1].fill = red_fill
                    
                    # 规则7: 如果是否在生效XBOM中的值为是，且其父级未冻结、未停止生产，则标红
                    xbom_value = str(data_row.get('是否在生效XBOM中', ''))
                    if '是，且其父级未冻结、未停止生产' in xbom_value:
                        if '是否在生效XBOM中' in col_indices:
                            excel_row[col_indices['是否在生效XBOM中']-1].fill = red_fill
                    
                    # 规则8: 如果是否被EBOM引用的值为是，且其父级未冻结、未停止生产，则标红
                    ebom_value = str(data_row.get('是否被EBOM引用', ''))
                    if '是，且其父级未冻结、未停止生产' in ebom_value:
                        if '是否被EBOM引用' in col_indices:
                            excel_row[col_indices['是否被EBOM引用']-1].fill = red_fill

                
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 保存带颜色的Excel文件
                wb.save(file_path)
                self.log_message(f"带颜色标记的结果已导出到: {file_path}")
                messagebox.showinfo("成功", "结果导出成功（已应用颜色标记）")
            except Exception as e:
                # 如果上色失败，回退到普通导出
                try:
                    self.result_df.to_excel(file_path, index=False)
                    self.log_message(f"结果已导出到: {file_path}（普通版本，上色功能出错: {str(e)}）")
                    messagebox.showinfo("部分成功", f"结果导出成功，但颜色标记未应用: {str(e)}")
                except Exception as e2:
                    self.log_message(f"导出时出错: {str(e2)}")
                    messagebox.showerror("错误", f"导出时出错: {str(e2)}")
                
    # 以下是原脚本中的核心函数（保持不变）
    def read_excel_files(self, folder_path):
        all_files = os.listdir(folder_path)
        excel_files = [file for file in all_files 
                      if file.endswith(('.xlsx', '.xls', '.XLSX', '.XLS')) 
                      and not file.startswith('~$')]
        
        if not excel_files:
            return None, None
            
        target_files = [file for file in excel_files if '物料冻结' in file]
        
        if not target_files:
            return None, excel_files
            
        base_file = target_files[0]
        base_file_path = os.path.join(folder_path, base_file)
        
        try:
            base_df = pd.read_excel(base_file_path)
            filtered_excel_files = [file for file in excel_files if file != base_file]
            return base_df, filtered_excel_files
        except Exception as e:
            self.log_message(f"读取基准文件失败: {str(e)}")
            return None, [file for file in excel_files if file != base_file]
            
    def calculate_inventory(self, result_df, current_df):
        current_df['非限制使用的库存'] = pd.to_numeric(current_df['非限制使用的库存'], errors='coerce').fillna(0)
        current_df['Trans./Tfr'] = pd.to_numeric(current_df['Trans./Tfr'], errors='coerce').fillna(0)
        inventory_mapping = {}
        for _, row in current_df.iterrows():
            key = row['物料编码']
            inventory_mapping[key] = inventory_mapping.get(key, 0) + row['非限制使用的库存'] + row['Trans./Tfr']
        #只有result_df的物料编码在inventory_mapping中，才对这一行进行map操作，但是不要修改result_df的大小
        for index,row in result_df.iterrows():
            if row['物料编码'] in inventory_mapping:
                result_df.at[index,'库存'] = inventory_mapping[row['物料编码']] + row['库存']
        return result_df
        
    def compare_material_codes_worklist(self, result_df, current_df):
        current_material_codes = set(current_df['物料编号'])
        for index,row in result_df.iterrows():
            if row['物料编码'] in current_material_codes:
                result_df.at[index,'是否有未结工单'] = '是'
        return result_df
        
    def compare_material_codes_buylist(self, result_df, current_df):
        current_material_codes = set(current_df['物料编码'])
        for index,row in result_df.iterrows():
            if row['物料编码'] in current_material_codes:
                result_df.at[index,'是否有未清采购订单'] = '是'
        return result_df
        
    def MBOM_mapping(self,result_df,current_df):
        MBOM_mapping = {}
        for index, row in current_df.iterrows():
            if row['是否在生效制造BOM中'] == '是，且其父级未冻结、未停止生产' and MBOM_mapping.get(row['子项物料编码'],0) == 0:
                MBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
            if row['是否在生效制造BOM中'] == '是，但其父级已冻结或停止生产' and MBOM_mapping.get(row['子项物料编码'],0) == 0:
                MBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'
            if row['是否在生效制造BOM中'] == '否' and MBOM_mapping.get(row['子项物料编码'],0) == 0:
                MBOM_mapping[row['子项物料编码']] = '否'
        result_df['是否在生效MBOM中'] = result_df['物料编码'].map(MBOM_mapping)

        return result_df
        
    def SBOM_mapping(self,result_df,current_df):
        SBOM_mapping = {}
        for index, row in current_df.iterrows():
            if row['是否在生效服务BOM中'] == '是，且其父级未冻结、未停止生产' and SBOM_mapping.get(row['子项物料编码'],0) == 0:
                SBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
                continue
            if row['是否在生效服务BOM中'] == '是，但其父级已冻结或停止生产' and SBOM_mapping.get(row['子项物料编码'],0) == 0:
                SBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'
                continue
            if row['是否在生效服务BOM中'] == '否' and SBOM_mapping.get(row['子项物料编码'],0) == 0:
                SBOM_mapping[row['子项物料编码']] = '否'
                continue
        result_df['是否在生效SBOM中'] = result_df['物料编码'].map(SBOM_mapping) 
        return result_df
        
    def EBOM_mapping(self,result_df,current_df):
        EBOM_mapping = {}
        for index, row in current_df.iterrows():
            if row['备注'] == '是，且其父级未冻结、未停止生产' and EBOM_mapping.get(row['子项物料编码'],0) == 0:
                EBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
                continue
            if row['备注'] == '是，但其父级已冻结或停止生产' and EBOM_mapping.get(row['子项物料编码'],0) == 0:
                EBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'
                continue
            if row['备注'] == '否' and EBOM_mapping.get(row['子项物料编码'],0) == 0:
                EBOM_mapping[row['子项物料编码']] = '否'
                continue
        result_df['是否被EBOM引用'] = result_df['物料编码'].map(EBOM_mapping) 
        return result_df
        
    def XBOM_mapping(self,result_df,current_df):
        XBOM_mapping = {}
        for index, row in current_df.iterrows():
            if row['是否在生效销售BOM中'] == '是，且其父级未冻结、未停止生产' and XBOM_mapping.get(row['子项物料编码'],0) == 0:
                XBOM_mapping[row['子项物料编码']] = '是，且其父级未冻结、未停止生产'
                continue
            if row['是否在生效销售BOM中'] == '是，但其父级已冻结或停止生产' and XBOM_mapping.get(row['子项物料编码'],0) == 0:
                XBOM_mapping[row['子项物料编码']] = '是，但其父级已冻结或停止生产'      
                continue
            if row['是否在生效销售BOM中'] == '否' and XBOM_mapping.get(row['子项物料编码'],0) == 0:
                XBOM_mapping[row['子项物料编码']] = '否'
                continue
        result_df['是否在生效XBOM中'] = result_df['物料编码'].map(XBOM_mapping) 
        return result_df
        
    def last_trade_date(self, result_df, current_df):
        last_trade_date_mapping = {}
        current_df['过帐日期'] = pd.to_datetime(current_df['过帐日期'], errors='coerce')
        current_df['凭证日期'] = pd.to_datetime(current_df['凭证日期'], errors='coerce')
        
        for _, row in current_df.iterrows():
            key = row['物料编号']
            dates = [d for d in [row['过帐日期'], row['凭证日期']] if pd.notna(d)]
            if dates:
                last_trade_date_mapping[key] = max(dates).strftime('%Y/%m/%d')
        for index,row in result_df.iterrows():
            if row['物料编码'] in last_trade_date_mapping:
                result_df.at[index,'最后一次交易日期'] = last_trade_date_mapping[row['物料编码']]
        return result_df
        
    def service_product(self, result_df, current_df):
        service_product_mapping = {}
        replace_product_mapping = {}
        replace_relation_mapping = {}
        current_df['替换物料'] = current_df['替换物料'].map(lambda x: '/' if x == 'nan' else str(int(float(x))))
        current_df['替换关系描述'] = current_df['替换关系描述'].map(lambda x: '/' if x == 'nan' else x)
        
        for _, row in current_df.iterrows():
            key = row['物料编码']
            service_product_mapping[key] = row['是否服务物料']
            if '替换物料' in current_df.columns:
                replace_product_mapping[row['物料编码']] = row['替换物料']
            if '替换关系描述' in current_df.columns:
                replace_relation_mapping[row['物料编码']] = row['替换关系描述']
                
        result_df['是否服务物料'] = result_df['物料编码'].map(service_product_mapping).fillna('没匹配到服务物料')
        
        if '替换物料' in current_df.columns:
            result_df['替换物料'] = result_df['物料编码'].map(replace_product_mapping).fillna('/')
        if '替换关系描述' in current_df.columns:
            result_df['替换关系描述'] = result_df['物料编码'].map(replace_relation_mapping).fillna('/')

        return result_df

if __name__ == "__main__":
    root = tk.Tk()
    app = MaterialFreezeGUI(root)
    root.mainloop()