"""
Excel数据核对工具 - 预处理前置优化版
功能：一对多数据核对、自动处理列名格式、主键拼接比对
优化：数据预处理前置、缓存机制、批量预处理、性能提升
"""

import pandas as pd
import os
import re
import json
import hashlib
import numpy as np
from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

# ---------------------- 数据预处理器 ----------------------
class DataPreprocessor:
    """数据预处理器 - 缓存机制"""
    
    def __init__(self, cache_dir="preprocess_cache"):
        self.cache_dir = cache_dir
        os.makedirs(cache_dir, exist_ok=True)
    
    def get_file_hash(self, file_path):
        """计算文件哈希值用于缓存识别"""
        try:
            hasher = hashlib.md5()
            with open(file_path, 'rb') as f:
                buf = f.read(65536)
                while len(buf) > 0:
                    hasher.update(buf)
                    buf = f.read(65536)
            return hasher.hexdigest()
        except Exception:
            return None
    
    def preprocess_and_cache(self, file_path, key_columns, sheet_name=0):
        """预处理并缓存数据"""
        if not os.path.exists(file_path):
            raise ValueError(f"文件不存在: {file_path}")
            
        file_hash = self.get_file_hash(file_path)
        if not file_hash:
            raise ValueError(f"无法读取文件: {file_path}")
            
        cache_file = os.path.join(self.cache_dir, f"{file_hash}.pkl")
        
        # 检查缓存
        if os.path.exists(cache_file):
            print(f"📁 使用缓存: {os.path.basename(file_path)}")
            return pd.read_pickle(cache_file)
        
        # 预处理数据
        print(f"🔄 预处理: {os.path.basename(file_path)}")
        df = self._read_and_preprocess(file_path, key_columns, sheet_name)
        
        # 保存缓存
        df.to_pickle(cache_file)
        return df
    
    def _read_and_preprocess(self, file_path, key_columns, sheet_name):
        """读取并预处理数据"""
        # 快速读取
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            dtype=str,
            keep_default_na=False,
            engine='openpyxl'
        )
        
        # 清洗列名
        df.columns = df.columns.str.strip().str.replace('\n', '')
        
        # 检查缺失列
        missing_keys = [col for col in key_columns if col not in df.columns]
        if missing_keys:
            raise ValueError(f"缺失主键列: {', '.join(missing_keys)}")
        
        # 批量预处理所有列
        for col in df.columns:
            if col in key_columns:
                continue  # 主键列跳过标准化
                
            col_lower = col.lower()
            
            # 根据列名类型进行相应预处理
            if any(word in col_lower for word in ['date', '日期', '时间']):
                df[col] = self._normalize_date_batch(df[col])
            elif any(word in col_lower for word in ['phone', '电话', '手机号', '手机']):
                df[col] = self._normalize_phone_batch(df[col])
            elif any(word in col_lower for word in ['金额', '数量', '价格', '数', '价', '金额']):
                df[col] = self._normalize_number_batch(df[col])
            else:
                df[col] = self._normalize_string_batch(df[col])
        
        # 生成组合主键
        df['_composite_key'] = self._generate_composite_key_vectorized(df, key_columns)
        
        return df
    
    def _normalize_date_batch(self, series):
        """批量日期标准化"""
        def parse_date(date_str):
            if pd.isna(date_str) or str(date_str).strip() in ['', '<空值>', 'nan', 'None', 'NaT']:
                return None
            
            formats = [
                '%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y',
                '%Y%m%d', '%Y.%m.%d', '%d-%b-%y',
                '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'
            ]
            
            date_str = str(date_str).strip()
            for fmt in formats:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.date().isoformat()
                except ValueError:
                    continue
            return date_str
        
        return series.apply(parse_date)
    
    def _normalize_phone_batch(self, series):
        """批量电话标准化"""
        def process_phone(phone):
            if pd.isna(phone) or str(phone).strip() in ['', '<空值>', 'nan', 'None', 'NaT']:
                return None
            
            cleaned = re.sub(r'\D', '', str(phone))
            
            if cleaned.startswith('86') and len(cleaned) > 11:
                return cleaned[2:]
            elif cleaned.startswith('0086'):
                return cleaned[4:]
            elif len(cleaned) == 11 and cleaned.startswith('1'):
                return cleaned
            return cleaned
        
        return series.apply(process_phone)
    
    def _normalize_number_batch(self, series):
        """批量数值标准化"""
        def process_number(num_str):
            if pd.isna(num_str) or str(num_str).strip() in ['', '<空值>', 'nan', 'None', 'NaT']:
                return None
            
            try:
                num = float(str(num_str).strip())
                return str(int(num)) if num.is_integer() else str(num)
            except ValueError:
                return str(num_str).strip()
        
        return series.apply(process_number)
    
    def _normalize_string_batch(self, series):
        """批量字符串标准化"""
        return series.astype(str).str.strip().replace(['nan', 'None', 'NaT'], '')
    
    def _generate_composite_key_vectorized(self, df, key_columns):
        """向量化生成组合主键"""
        key_parts = []
        for col in key_columns:
            if col in df.columns:
                series = df[col].astype(str).str.strip()
                series = series.replace(['', 'nan', 'None', 'NaT'], '<空值>')
                key_parts.append(series)
        
        if key_parts:
            composite_key = key_parts[0]
            for part in key_parts[1:]:
                composite_key = composite_key + '_' + part
            return composite_key
        return pd.Series([''] * len(df))

# ---------------------- 高性能比对函数（预处理版本） ----------------------
def compare_preprocessed_datasets(source_df, target_df):
    """预处理后的数据比对函数"""
    
    # 处理标准格式
    if len(target_df) > 0 and str(target_df.iloc[0, 0]).strip() == '元数据标准名称':
        print('==========标准模板处理===========')
        target_df = target_df.iloc[2:]  # 跳过前两行
        if len(target_df.columns) > 1:
            target_df = target_df.drop(target_df.columns[0], axis=1)
        target_df = target_df.reset_index(drop=True)
    
    # 直接使用预处理后的组合主键
    source_key_map = dict(zip(source_df['_composite_key'], source_df.index))
    target_key_map = dict(zip(target_df['_composite_key'], target_df.index))
    
    source_keys = set(source_key_map.keys())
    target_keys = set(target_key_map.keys())
    common_keys = source_keys & target_keys
    
    # 获取共同字段（排除预处理添加的列）
    common_columns = [col for col in set(source_df.columns) & set(target_df.columns) 
                     if col not in ['_composite_key']]
    
    results = []
    
    # 批量处理仅存在于源数据的主键
    for key in source_keys - target_keys:
        idx = source_key_map[key]
        row_data = {'主键状态': '仅存在于源文件', '组合主键': key}
        for col in common_columns:
            row_data[f"源_{col}"] = source_df.at[idx, col]
            row_data[f"目标_{col}"] = ""
        results.append(row_data)
    
    # 批量处理仅存在于目标数据的主键
    for key in target_keys - source_keys:
        idx = target_key_map[key]
        row_data = {'主键状态': '仅存在于目标文件', '组合主键': key}
        for col in common_columns:
            row_data[f"源_{col}"] = ""
            row_data[f"目标_{col}"] = target_df.at[idx, col]
        results.append(row_data)
    
    # 批量处理共同主键的数据差异（使用预处理后的值直接比较）
    for key in common_keys:
        src_idx = source_key_map[key]
        tgt_idx = target_key_map[key]
        
        row_data = {'主键状态': '数据一致', '组合主键': key}
        diff_details = {}
        
        for col in common_columns:
            src_val = source_df.at[src_idx, col]
            tgt_val = target_df.at[tgt_idx, col]
            
            # 预处理后的值直接比较
            src_empty = pd.isna(src_val) or str(src_val).strip() in ['', '<空值>']
            tgt_empty = pd.isna(tgt_val) or str(tgt_val).strip() in ['', '<空值>']
            
            row_data[f"源_{col}"] = src_val
            row_data[f"目标_{col}"] = tgt_val
            
            if not (src_empty and tgt_empty) and str(src_val) != str(tgt_val):
                diff_details[col] = {'源值': src_val, '目标值': tgt_val}
        
        if diff_details:
            row_data['主键状态'] = f"发现{len(diff_details)}处差异"
            row_data['差异详情'] = str(diff_details)
            row_data['差异列名'] = ",".join(diff_details.keys())
        else:
            row_data['差异详情'] = None
        
        results.append(row_data)
    
    return pd.DataFrame(results)

# ---------------------- 报告生成函数 ----------------------
def generate_detailed_report_optimized(result_df, output_path):
    """优化的报告生成函数"""
    
    # 重新排列列的顺序：主键状态、组合主键、源文件列、目标文件列、差异详情
    all_columns = result_df.columns.tolist()
    
    # 提取固定列
    status_col = '主键状态'
    key_col = '组合主键'
    diff_col = '差异详情'
    
    # 提取源文件列和目标文件列
    source_cols = [col for col in all_columns if col.startswith('源_')]
    target_cols = [col for col in all_columns if col.startswith('目标_')]
    
    # 构建新的列顺序
    new_column_order = [status_col, key_col]
    new_column_order.extend(source_cols)
    new_column_order.extend(target_cols)
    
    # 添加其他列（差异详情等）
    other_cols = [col for col in all_columns if col not in new_column_order and col != status_col and col != key_col]
    new_column_order.extend(other_cols)
    
    # 重新排序DataFrame
    result_df = result_df[new_column_order]
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='核对结果')
        
        workbook = writer.book
        worksheet = writer.sheets['核对结果']
        
        # 定义颜色
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # 获取列索引
        status_col = 1  # A列
        diff_col = len(result_df.columns)  # 最后一列
        
        # 批量获取需要标记的行
        yellow_rows = result_df[result_df['主键状态'].isin(['仅存在于源文件', '仅存在于目标文件'])].index + 2
        red_rows = result_df[result_df['主键状态'].str.contains('差异', na=False)].index + 2
        
        # 批量应用样式
        for row_idx in yellow_rows:
            for col_idx in range(1, len(result_df.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
        
        for row_idx in red_rows:
            worksheet.cell(row=row_idx, column=status_col).fill = red_fill
            worksheet.cell(row=row_idx, column=diff_col).fill = red_fill
            
            # 标记差异列
            if row_idx - 2 < len(result_df):
                diff_details = result_df.iloc[row_idx - 2]['差异详情']
                if diff_details and isinstance(diff_details, str):
                    try:
                        import ast
                        diff_dict = ast.literal_eval(diff_details)
                        for col_name in diff_dict.keys():
                            for col_idx, col in enumerate(result_df.columns):
                                if col_name in col:
                                    worksheet.cell(row=row_idx, column=col_idx + 1).fill = red_fill
                    except:
                        pass
        
        # 优化列宽设置
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        
        worksheet.freeze_panes = 'A2'

# ---------------------- 批量处理主程序（预处理前置版本） ----------------------
def batch_compare_with_preprocess(source_path, target_paths, key_columns, output_dir):
    """使用预处理缓存的批量比对主程序"""
    
    os.makedirs(output_dir, exist_ok=True)
    preprocessor = DataPreprocessor()
    
    try:
        print(f"\n{'=' * 50} 开始预处理 {'=' * 50}")
        start_total = datetime.now()
        
        # 预处理源数据
        print(f"[{datetime.now().strftime('%H:%M:%S')}] 预处理源文件: {os.path.basename(source_path)}")
        source_df = preprocessor.preprocess_and_cache(source_path, key_columns)
        print(f"源文件预处理完成，记录数: {len(source_df):,}")
        
        # 预处理所有目标数据
        target_dfs = []
        for i, target_path in enumerate(target_paths, 1):
            try:
                target_df = preprocessor.preprocess_and_cache(target_path, key_columns)
                target_dfs.append((target_path, target_df))
                print(f"目标文件 {i}/{len(target_paths)} 预处理完成，记录数: {len(target_df):,}")
            except Exception as e:
                print(f"❌ 预处理失败: {os.path.basename(target_path)} - {str(e)}")
        
        print(f"\n{'=' * 50} 开始比对 {'=' * 50}")
        
        # 使用预处理后的数据进行比对
        for target_path, target_df in target_dfs:
            start_time = datetime.now()
            target_name = os.path.basename(target_path)
            print(f"\n[{start_time.strftime('%H:%M:%S')}] 比对: {target_name}")
            
            try:
                # 高性能比对（无需再次预处理）
                result_df = compare_preprocessed_datasets(source_df, target_df)
                
                # 生成报告
                report_name = f"比对报告_{os.path.splitext(os.path.basename(source_path))[0]}_vs_{os.path.splitext(target_name)[0]}.xlsx"
                output_path = os.path.join(output_dir, report_name)
                generate_detailed_report_optimized(result_df, output_path)
                
                # 统计信息
                duration = (datetime.now() - start_time).total_seconds()
                total_duration = (datetime.now() - start_total).total_seconds()
                
                stats = f"""
✅ 完成比对 ({duration:.2f}秒)
📊 差异统计：
   • 仅存在于源文件: {len(result_df[result_df['主键状态'] == '仅存在于源文件']):,}
   • 仅存在于目标文件: {len(result_df[result_df['主键状态'] == '仅存在于目标文件']):,}
   • 存在数据差异: {len(result_df[result_df['主键状态'].str.contains('差异', na=False)]):,}
   • 总计记录: {len(result_df):,}
💾 报告已保存至: {output_path}
⏱️ 累计耗时: {total_duration:.2f}秒
"""
                print(stats)
                
            except Exception as e:
                print(f"❌ 比对失败: {str(e)}")
        
        total_time = (datetime.now() - start_total).total_seconds()
        print(f"\n🎉 全部处理完成！总耗时: {total_time:.2f}秒")
        
    except Exception as e:
        print(f"\n❌ 严重错误: {str(e)}")
    finally:
        print(f"{'=' * 50} 处理完成 {'=' * 50}")

# ---------------------- 兼容旧接口 ----------------------
def batch_compare_optimized(source_path, target_paths, key_columns, output_dir):
    """兼容旧接口，使用预处理前置优化"""
    return batch_compare_with_preprocess(source_path, target_paths, key_columns, output_dir)

if __name__ == "__main__":
    # ===================== 配置区 =====================
    SOURCE_FILE = r"C:\Users\zhangbon\Desktop\案例\核对数据集\MDM中销售BOM.XLSX"
    TARGET_FILES = [
        r"C:\Users\zhangbon\Desktop\案例\核对数据集\ERP销售BOM.XLSX"
    ]
    KEY_COLUMNS = ["工厂", "物料编码", "组件"]
    OUTPUT_DIR = r"C:\Users\zhangbon\Desktop\案例\核对数据集\输出报告"
    # ================================================
    
    batch_compare_with_preprocess(SOURCE_FILE, TARGET_FILES, KEY_COLUMNS, OUTPUT_DIR)