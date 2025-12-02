"""
Excel数据核对工具 - 简化版GUI界面
功能：一对多数据核对、自动处理列名格式、主键拼接比对
简化：移除缓存机制，优化数值转换，保持列名顺序
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
from openpyxl.styles import PatternFill
from datetime import datetime
import threading
import warnings
warnings.filterwarnings('ignore')

def normalize_number(value):
    """数值标准化 - 修复5.0等数值转换问题"""
    if pd.isna(value) or str(value).strip() in ['', '<空值>', 'nan', 'None', 'NaT']:
        return None
    
    try:
        # 先尝试转换为浮点数
        num = float(str(value).strip())
        # 如果是整数（如5.0），转换为整数再转字符串
        if num.is_integer():
            return str(int(num))
        else:
            return str(num)
    except (ValueError, TypeError):
        # 如果不是数值，返回原字符串（去除空格）
        return str(value).strip()

def normalize_string(value):
    """字符串标准化"""
    if pd.isna(value):
        return ''
    return str(value).strip()

def read_and_preprocess(file_path, key_columns, sheet_name=0):
    """读取并预处理数据 - 简化版"""
    print(f"📖 读取文件: {os.path.basename(file_path)}")
    
    # 读取Excel文件
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        dtype=str,
        keep_default_na=False,
        engine='openpyxl'
    )
    
    # 清洗列名
    df.columns = df.columns.str.strip().str.replace('\n', '')
    
    # 检查缺失列
    missing_keys = [col for col in key_columns if col not in df.columns]
    if missing_keys:
        raise ValueError(f"缺失主键列: {', '.join(missing_keys)}")
    
    # 预处理所有列
    for col in df.columns:
        df[col] = df[col].apply(normalize_string)
        df[col] = df[col].apply(normalize_number)
    
    # 生成组合主键
    key_parts = []
    for col in key_columns:
        if col in df.columns:
            # 标准化空值处理
            series = df[col].fillna('').astype(str).str.strip()
            series = series.replace(['', 'nan', 'None', 'NaT'], '<空值>')
            key_parts.append(series)
    
    if key_parts:
        df['_composite_key'] = key_parts[0]
        for part in key_parts[1:]:
            df['_composite_key'] = df['_composite_key'] + '_' + part
    else:
        df['_composite_key'] = ''
    
    return df

def compare_datasets(source_df, target_df, original_columns):
    """数据比对函数 - 只核对源文件和目标文件共有的字段"""
    
    # # 处理标准格式（跳过前两行）
    # if len(target_df) > 0 and str(target_df.iloc[0, 0]).strip() == '元数据标准名称':
    #     print('==========标准模板处理===========')
    #     target_df = target_df.iloc[2:].reset_index(drop=True)
    
    # 获取主键映射
    source_key_map = dict(zip(source_df['_composite_key'], source_df.index))
    target_key_map = dict(zip(target_df['_composite_key'], target_df.index))
    
    source_keys = set(source_key_map.keys())
    target_keys = set(target_key_map.keys())
    common_keys = source_keys & target_keys
    
    # 获取源文件和目标文件共有的字段
    source_columns = set(source_df.columns) - {'_composite_key'}
    target_columns = set(target_df.columns) - {'_composite_key'}
    common_columns = source_columns & target_columns
    
    # 按照原始列顺序过滤共有字段
    columns_to_compare = [col for col in original_columns if col in common_columns]
    
    results = []
    
    # 处理仅存在于源数据的主键
    for key in source_keys - target_keys:
        idx = source_key_map[key]
        row_data = {'主键状态': '仅存在于源文件', '组合主键': key}
        
        # 按照原始列顺序添加数据，但只保留共有的字段
        for col in original_columns:
            if col in source_df.columns and col in common_columns:
                row_data[f"源_{col}"] = source_df.at[idx, col]
                row_data[f"目标_{col}"] = ""
        results.append(row_data)
    
    # 处理仅存在于目标数据的主键
    for key in target_keys - source_keys:
        idx = target_key_map[key]
        row_data = {'主键状态': '仅存在于目标文件', '组合主键': key}
        
        # 按照原始列顺序添加数据，但只保留共有的字段
        for col in original_columns:
            if col in target_df.columns and col in common_columns:
                row_data[f"源_{col}"] = ""
                row_data[f"目标_{col}"] = target_df.at[idx, col]
        results.append(row_data)
    
    # 处理共同主键的数据差异
    for key in common_keys:
        src_idx = source_key_map[key]
        tgt_idx = target_key_map[key]
        
        row_data = {'主键状态': '数据一致', '组合主键': key}
        diff_count = 0
        diff_details = {}
        
        # 只比较源文件和目标文件共有的字段
        for col in columns_to_compare:
            src_val = source_df.at[src_idx, col] if col in source_df.columns else None
            tgt_val = target_df.at[tgt_idx, col] if col in target_df.columns else None
            
            # 标准化空值比较
            src_empty = pd.isna(src_val) or str(src_val).strip() in ['', '<空值>']
            tgt_empty = pd.isna(tgt_val) or str(tgt_val).strip() in ['', '<空值>']
            
            row_data[f"源_{col}"] = src_val if src_val is not None else ""
            row_data[f"目标_{col}"] = tgt_val if tgt_val is not None else ""
            
            # 检查差异（两个都非空且不相等）
            if not (src_empty and tgt_empty) and str(src_val) != str(tgt_val):
                diff_count += 1
                diff_details[col] = {'源值': src_val, '目标值': tgt_val}
        
        if diff_count > 0:
            row_data['主键状态'] = f"发现{diff_count}处差异"
            row_data['差异详情'] = str(diff_details)
            row_data['差异列名'] = ",".join(diff_details.keys())
        else:
            row_data['差异详情'] = None
            row_data['差异列名'] = None
        
        results.append(row_data)
    
    return pd.DataFrame(results)

def generate_report(result_df, output_path, original_columns):
    """生成报告 - 保持列顺序，并对差异列进行标红"""
    
    # 构建列顺序：固定列 + 按照原始顺序的源列和目标列
    column_order = ['主键状态', '组合主键']
    
    # 按照原始列顺序添加源列和目标列
    for col in original_columns:
        column_order.append(f"源_{col}")
        column_order.append(f"目标_{col}")
    
    # 添加差异详情列
    column_order.extend(['差异详情', '差异列名'])
    
    # 重新排序（只保留存在的列）
    existing_columns = [col for col in column_order if col in result_df.columns]
    result_df = result_df[existing_columns]
    
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        result_df.to_excel(writer, index=False, sheet_name='核对结果')
        
        workbook = writer.book
        worksheet = writer.sheets['核对结果']
        
        # 定义颜色
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # 批量标记颜色
        for row_idx in range(2, len(result_df) + 2):
            status = worksheet.cell(row=row_idx, column=1).value
            
            if status in ['仅存在于源文件', '仅存在于目标文件']:
                # 黄色标记整行
                for col_idx in range(1, len(result_df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
            
            elif status and '差异' in str(status):
                # 红色标记状态列
                worksheet.cell(row=row_idx, column=1).fill = red_fill  # 主键状态
                
                # 获取该行的差异列名
                diff_cols_cell = worksheet.cell(row=row_idx, column=result_df.columns.get_loc('差异列名') + 1)
                if diff_cols_cell.value:
                    diff_cols = diff_cols_cell.value.split(',')
                    
                    # 对每个差异列进行标红
                    for col_name in diff_cols:
                        # 查找源列和目标列的索引位置
                        source_col_name = f"源_{col_name}"
                        target_col_name = f"目标_{col_name}"
                        
                        if source_col_name in result_df.columns:
                            source_col_idx = result_df.columns.get_loc(source_col_name) + 1
                            worksheet.cell(row=row_idx, column=source_col_idx).fill = red_fill
                        
                        if target_col_name in result_df.columns:
                            target_col_idx = result_df.columns.get_loc(target_col_name) + 1
                            worksheet.cell(row=row_idx, column=target_col_idx).fill = red_fill
                
                # 红色标记差异详情列
                if '差异详情' in result_df.columns:
                    diff_col_idx = result_df.columns.get_loc('差异详情') + 1
                    worksheet.cell(row=row_idx, column=diff_col_idx).fill = red_fill
        
        # 设置列宽
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        
        worksheet.freeze_panes = 'A2'

class DataComparisonApp:
    def __init__(self, root):
        self.root = root
        self.root.title("数据核对工具 - 简化版")
        self.root.geometry("800x600")
        
        # 文件路径变量
        self.source_path = tk.StringVar()
        self.target_paths = []
        self.key_columns = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.join(os.getcwd(), "核对报告"))
        
        self.setup_ui()
    
    def setup_ui(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置区域
        config_frame = ttk.LabelFrame(main_frame, text="配置选项", padding="10")
        config_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 源文件选择
        ttk.Label(config_frame, text="源文件:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.source_path, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(config_frame, text="浏览", command=self.select_source_file).grid(row=0, column=2, padx=5)
        
        # 目标文件选择
        ttk.Label(config_frame, text="目标文件:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.target_listbox = tk.Listbox(config_frame, height=3, width=50)
        self.target_listbox.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5)
        
        target_btn_frame = ttk.Frame(config_frame)
        target_btn_frame.grid(row=1, column=2, padx=5)
        ttk.Button(target_btn_frame, text="添加", command=self.add_target_file).pack(pady=2)
        ttk.Button(target_btn_frame, text="删除", command=self.remove_target_file).pack(pady=2)
        
        # 主键列设置
        ttk.Label(config_frame, text="主键列名:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.key_columns, width=50).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Label(config_frame, text="(用逗号分隔)").grid(row=2, column=2, sticky=tk.W)
        
        # 源文件列名显示区域
        ttk.Label(config_frame, text="源文件列名:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.source_columns_text = tk.Text(config_frame, height=3, width=50, wrap=tk.WORD)
        self.source_columns_text.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=5)
        source_columns_scrollbar = ttk.Scrollbar(config_frame, orient=tk.VERTICAL, command=self.source_columns_text.yview)
        self.source_columns_text.configure(yscrollcommand=source_columns_scrollbar.set)
        source_columns_scrollbar.grid(row=3, column=2, sticky=(tk.N, tk.S))
        
        # 目标文件列名显示区域
        ttk.Label(config_frame, text="目标文件列名:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.target_columns_text = tk.Text(config_frame, height=3, width=50, wrap=tk.WORD)
        self.target_columns_text.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=5)
        target_columns_scrollbar = ttk.Scrollbar(config_frame, orient=tk.VERTICAL, command=self.target_columns_text.yview)
        self.target_columns_text.configure(yscrollcommand=target_columns_scrollbar.set)
        target_columns_scrollbar.grid(row=4, column=2, sticky=(tk.N, tk.S))
        
        # 输出目录
        ttk.Label(config_frame, text="输出目录:").grid(row=5, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.output_dir, width=50).grid(row=5, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(config_frame, text="浏览", command=self.select_output_dir).grid(row=5, column=2, padx=5)
        
        # 按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=1, column=0, columnspan=2, pady=10)
        
        ttk.Button(btn_frame, text="开始核对", command=self.start_comparison, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="清空", command=self.clear_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="退出", command=self.root.quit).pack(side=tk.LEFT, padx=5)
        
        # 日志区域
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="10")
        log_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 日志文本框和滚动条
        self.log_text = tk.Text(log_frame, height=15, width=80, wrap=tk.WORD)
        log_scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # 配置网格权重
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        config_frame.columnconfigure(1, weight=1)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # 设置样式
        style = ttk.Style()
        style.configure("Accent.TButton", foreground="white", background="#0078d4")
    
    def log_message(self, message):
        """添加日志消息"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def select_source_file(self):
        """选择源文件"""
        filename = filedialog.askopenfilename(
            title="选择源文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if filename:
            self.source_path.set(filename)
            self.log_message(f"已选择源文件: {os.path.basename(filename)}")
            
            # 显示源文件列名
            columns = self.get_file_columns(filename)
            if columns:
                self.source_columns_text.delete(1.0, tk.END)
                self.source_columns_text.insert(tk.END, f"源文件列名 ({len(columns)}列):\n")
                self.source_columns_text.insert(tk.END, ", ".join(columns))
                self.log_message(f"源文件包含 {len(columns)} 列: {', '.join(columns[:5])}{'...' if len(columns) > 5 else ''}")
            else:
                self.source_columns_text.delete(1.0, tk.END)
                self.source_columns_text.insert(tk.END, "无法读取文件列名")
    
    def add_target_file(self):
        """添加目标文件"""
        filenames = filedialog.askopenfilenames(
            title="选择目标文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if filenames:
            # 清空目标文件列名显示区域，准备显示新信息
            self.target_columns_text.delete(1.0, tk.END)
            
            # 处理每个文件
            for i, filename in enumerate(filenames, 1):
                if filename not in self.target_paths:
                    self.target_paths.append(filename)
                    self.target_listbox.insert(tk.END, os.path.basename(filename))
                    self.log_message(f"已添加目标文件: {os.path.basename(filename)}")
            
            # 更新目标文件列名显示
            self.update_target_columns_display()
    
    def remove_target_file(self):
        """删除选中的目标文件"""
        selection = self.target_listbox.curselection()
        if selection:
            index = selection[0]
            filename = self.target_paths.pop(index)
            self.target_listbox.delete(index)
            self.log_message(f"已删除目标文件: {os.path.basename(filename)}")
            
            # 更新目标文件列名显示
            self.update_target_columns_display()
    
    def select_output_dir(self):
        """选择输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log_message(f"已选择输出目录: {directory}")
    
    def clear_all(self):
        """清空所有设置"""
        self.source_path.set("")
        self.target_paths.clear()
        self.target_listbox.delete(0, tk.END)
        self.key_columns.set("")
        self.output_dir.set(os.path.join(os.getcwd(), "核对报告"))
        self.source_columns_text.delete(1.0, tk.END)
        self.target_columns_text.delete(1.0, tk.END)
        self.log_text.delete(1.0, tk.END)
        self.log_message("已清空所有设置")
        
        # 初始化显示提示
        self.source_columns_text.insert(tk.END, "请选择源文件以显示列名")
        self.target_columns_text.insert(tk.END, "请选择目标文件以显示列名")
    
    def get_file_columns(self, file_path):
        """读取Excel文件的列名"""
        try:
            # 读取Excel文件的第一行（列名）
            df = pd.read_excel(file_path, nrows=1, engine='openpyxl')
            columns = df.columns.tolist()
            return columns
        except Exception as e:
            self.log_message(f"❌ 读取文件列名失败: {str(e)}")
            return []
    
    def update_target_columns_display(self):
        """更新目标文件列名显示"""
        self.target_columns_text.delete(1.0, tk.END)
        
        # 如果有目标文件，显示目标文件信息
        if self.target_paths:
            if len(self.target_paths) == 1:
                # 单个目标文件，显示完整列名
                filename = self.target_paths[0]
                columns = self.get_file_columns(filename)
                if columns:
                    self.target_columns_text.insert(tk.END, f"目标文件列名 ({len(columns)}列):\n")
                    self.target_columns_text.insert(tk.END, ", ".join(columns))
                    self.log_message(f"目标文件包含 {len(columns)} 列")
                else:
                    self.target_columns_text.insert(tk.END, "无法读取目标文件列名")
            else:
                # 多个目标文件，显示汇总信息
                self.target_columns_text.insert(tk.END, f"已选择 {len(self.target_paths)} 个目标文件\n\n")
                for i, filename in enumerate(self.target_paths, 1):
                    columns = self.get_file_columns(filename)
                    if columns:
                        file_info = f"{i}. {os.path.basename(filename)} ({len(columns)}列):\n"
                        file_info += f"   列名: {', '.join(columns[:3])}{'...' if len(columns) > 3 else ''}\n\n"
                        self.target_columns_text.insert(tk.END, file_info)
                        self.log_message(f"目标文件 {os.path.basename(filename)} 包含 {len(columns)} 列")
                    else:
                        self.target_columns_text.insert(tk.END, f"{i}. {os.path.basename(filename)}: 无法读取列名\n\n")
        else:
            # 没有目标文件时显示提示
            self.target_columns_text.insert(tk.END, "请选择目标文件以显示列名")
    
    def validate_inputs(self):
        """验证输入"""
        if not self.source_path.get():
            messagebox.showerror("错误", "请选择源文件")
            return False
        
        if not self.target_paths:
            messagebox.showerror("错误", "请添加至少一个目标文件")
            return False
        
        if not self.key_columns.get():
            messagebox.showerror("错误", "请设置主键列名")
            return False
        
        if not self.output_dir.get():
            messagebox.showerror("错误", "请选择输出目录")
            return False
        
        return True
    
    def start_comparison(self):
        """开始核对"""
        if not self.validate_inputs():
            return
        
        # 禁用开始按钮
        self.root.config(cursor="watch")
        
        # 在新线程中运行比对
        thread = threading.Thread(target=self.run_comparison)
        thread.daemon = True
        thread.start()
    
    def run_comparison(self):
        """运行比对"""
        try:
            source_path = self.source_path.get()
            key_columns = [col.strip() for col in self.key_columns.get().split(',')]
            output_dir = self.output_dir.get()
            
            self.log_message("=" * 50)
            self.log_message("开始数据核对...")
            self.log_message(f"源文件: {os.path.basename(source_path)}")
            self.log_message(f"目标文件数量: {len(self.target_paths)}")
            self.log_message(f"主键列: {', '.join(key_columns)}")
            self.log_message(f"输出目录: {output_dir}")
            
            # 运行简化版的比对
            batch_compare_simple(source_path, self.target_paths, key_columns, output_dir)
            
            self.log_message("✅ 数据核对完成！")
            self.log_message("=" * 50)
            
            # 显示完成消息
            self.root.after(0, lambda: messagebox.showinfo("完成", "数据核对已完成！"))
            
        except Exception as e:
            error_msg = f"❌ 处理失败: {str(e)}"
            self.log_message(error_msg)
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        
        finally:
            # 恢复光标
            self.root.after(0, lambda: self.root.config(cursor=""))

def batch_compare_simple(source_path, target_paths, key_columns, output_dir):
    """简化的批量比对主程序"""
    
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        print(f"\n{'=' * 50} 开始处理 {'=' * 50}")
        start_total = datetime.now()
        
        # 读取源文件并获取原始列顺序
        print(f"[{datetime.now().strftime('%H:%M:%S')}] 读取源文件: {os.path.basename(source_path)}")
        source_df = read_and_preprocess(source_path, key_columns)
        original_columns = [col for col in source_df.columns if col not in ['_composite_key']]
        print(f"源文件读取完成，记录数: {len(source_df):,}")
        
        # 处理每个目标文件
        for i, target_path in enumerate(target_paths, 1):
            start_time = datetime.now()
            target_name = os.path.basename(target_path)
            print(f"\n[{start_time.strftime('%H:%M:%S')}] 处理: {target_name}")
            
            try:
                # 读取目标文件
                target_df = read_and_preprocess(target_path, key_columns)
                print(f"目标文件读取完成，记录数: {len(target_df):,}")
                
                # 比对数据
                result_df = compare_datasets(source_df, target_df, original_columns)
                
                # 生成报告
                report_name = f"比对报告_{os.path.splitext(os.path.basename(source_path))[0]}_vs_{os.path.splitext(target_name)[0]}.xlsx"
                output_path = os.path.join(output_dir, report_name)
                generate_report(result_df, output_path, original_columns)
                
                # 清理目标文件数据缓存
                del target_df
                del result_df
                import gc
                gc.collect()
                print("✅ 目标文件数据缓存已清理")
                
                # 统计信息
                duration = (datetime.now() - start_time).total_seconds()
                
                stats = f"""
✅ 完成比对 ({duration:.2f}秒)
📊 差异统计：
   • 仅存在于源文件: {len(result_df[result_df['主键状态'] == '仅存在于源文件']):,}
   • 仅存在于目标文件: {len(result_df[result_df['主键状态'] == '仅存在于目标文件']):,}
   • 存在数据差异: {len(result_df[result_df['主键状态'].str.contains('差异', na=False)]):,}
   • 总计记录: {len(result_df):,}
💾 报告已保存至: {output_path}
"""
                print(stats)
                
            except Exception as e:
                print(f"❌ 处理失败: {str(e)}")
        
        # 清理源文件数据缓存
        del source_df
        import gc
        gc.collect()
        print("✅ 源文件数据缓存已清理")
        
        total_time = (datetime.now() - start_total).total_seconds()
        print(f"\n🎉 全部处理完成！总耗时: {total_time:.2f}秒")
        
    except Exception as e:
        print(f"\n❌ 严重错误: {str(e)}")
    finally:
        # 最终清理
        import gc
        gc.collect()
        print(f"{'=' * 50} 处理完成 {'=' * 50}")

if __name__ == "__main__":
    # 创建主窗口
    root = tk.Tk()
    
    # 设置窗口图标和样式（如果存在）
    try:
        root.iconbitmap(default='excel.ico')
    except:
        pass
    
    # 创建应用
    app = DataComparisonApp(root)
    
    # 运行主循环
    root.mainloop()