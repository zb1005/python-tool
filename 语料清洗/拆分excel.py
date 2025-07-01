import os
import sys
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import shutil
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QWidget, 
                            QListWidget, QPushButton, QLabel, QFileDialog, 
                            QMessageBox, QProgressBar, QHBoxLayout)
from PyQt5.QtCore import Qt

class ExcelSplitterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel拆分工具 - 专业版")
        self.setGeometry(100, 100, 700, 500)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QListWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            QLabel {
                font-size: 14px;
            }
        """)
        
        # 主界面组件
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout()
        
        # 标题
        self.title_label = QLabel("Excel工作表拆分工具")
        self.title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.title_label.setAlignment(Qt.AlignCenter)
        
        # 文件列表
        self.file_list = QListWidget()
        self.file_list.setAcceptDrops(True)
        self.file_list.setStyleSheet("font-size: 12px;")
        
        # 按钮布局
        self.button_layout = QHBoxLayout()
        self.add_button = QPushButton("添加文件")
        self.add_button.clicked.connect(self.add_files)
        self.clear_button = QPushButton("清空列表")
        self.clear_button.clicked.connect(self.clear_files)
        self.split_button = QPushButton("开始拆分")
        self.split_button.clicked.connect(self.split_files)
        self.split_button.setStyleSheet("background-color: #2196F3;")
        
        self.button_layout.addWidget(self.add_button)
        self.button_layout.addWidget(self.clear_button)
        self.button_layout.addWidget(self.split_button)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        
        # 添加组件到布局
        self.layout.addWidget(self.title_label)
        self.layout.addWidget(QLabel("拖拽Excel文件到这里或点击添加按钮:"))
        self.layout.addWidget(self.file_list)
        self.layout.addLayout(self.button_layout)
        self.layout.addWidget(self.progress_bar)
        
        self.central_widget.setLayout(self.layout)
    
    def dragEnterEvent(self, event):
        """接受拖拽事件"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()
    
    def dropEvent(self, event):
        """处理拖放的文件"""
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.CopyAction)
            event.accept()
            
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if file_path.lower().endswith('.xlsx'):
                    # 避免重复添加
                    if not any(self.file_list.item(i).text() == file_path 
                             for i in range(self.file_list.count())):
                        self.file_list.addItem(file_path)
        else:
            event.ignore()
    
    def add_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx)"
        )
        for file in files:
            self.file_list.addItem(file)
    
    def clear_files(self):
        self.file_list.clear()
    
    def split_files(self):
        if self.file_list.count() == 0:
            QMessageBox.warning(self, "警告", "请先添加Excel文件!")
            return
        
        output_dir = QFileDialog.getExistingDirectory(
            self, "选择输出文件夹"
        )
        if not output_dir:
            return
        
        total_files = self.file_list.count()
        self.progress_bar.setRange(0, total_files)
        
        for i in range(total_files):
            input_file = self.file_list.item(i).text()
            try:
                split_excel_with_images(input_file, output_dir)
                self.progress_bar.setValue(i + 1)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"处理文件 {os.path.basename(input_file)} 时出错:\n{str(e)}")
                continue
        
        self.file_list.clear()
        self.progress_bar.setValue(0)
        QMessageBox.information(self, "完成", f"成功拆分 {total_files} 个Excel文件!")

def split_excel_with_images(input_file, output_folder):
    # 获取原始文件名(不带扩展名)
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    
    # 创建输出文件夹
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # 加载原始Excel文件
    wb = load_workbook(input_file)
    
    # 获取所有工作表名称
    sheet_names = wb.sheetnames
    
    # 临时文件夹用于存储图片
    temp_img_folder = os.path.join(output_folder, "_temp_images")
    if not os.path.exists(temp_img_folder):
        os.makedirs(temp_img_folder)
    
    # 遍历每个工作表
    for sheet_name in sheet_names:
        # 创建新工作簿
        new_wb = load_workbook(input_file)
        
        # 删除不需要的工作表
        for sheet in new_wb.sheetnames:
            if sheet != sheet_name:
                # 确保不是最后一个工作表
                if len(new_wb.sheetnames) > 1:
                    del new_wb[sheet]
        
        # 处理图片 - 先清空原有图片
        ws = new_wb[sheet_name]
        ws._images = []  # 清空现有图片避免重复
        
        # 从原始工作簿获取图片
        original_ws = wb[sheet_name]
        for idx, img in enumerate(original_ws._images):
            # 保存图片到临时文件夹
            img_path = os.path.join(temp_img_folder, f"{sheet_name}_{idx}.png")
            with open(img_path, "wb") as f:
                f.write(img._data())
            
            # 重新插入图片
            new_img = Image(img_path)
            ws.add_image(new_img)
        
        # 保存新文件
        # 修改保存文件名格式
        output_file = os.path.join(output_folder, f"{base_name}_{sheet_name}.xlsx")
        new_wb.save(output_file)
        print(f"已保存: {output_file}")
    
    # 删除临时图片文件夹
    shutil.rmtree(temp_img_folder)
    print("处理完成!")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelSplitterApp()
    window.show()
    sys.exit(app.exec_())
    # 使用示例
    input_excel = input("请输入要拆分的Excel文件路径: ")
    output_dir = input("请输入输出文件夹路径: ")
    
    split_excel_with_images(input_excel, output_dir)