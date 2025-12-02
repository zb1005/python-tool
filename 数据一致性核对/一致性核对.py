import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
from openpyxl.styles import PatternFill
from datetime import datetime
import threading
import warnings
warnings.filterwarnings('ignore')


def normalize_number(value):
    """数值标准化 - 修复5.0等数值转换问题"""
    if pd.isna(value) or str(value).strip() in ['', '<空值>', 'nan', 'None', 'NaT']:
        return None
    
    try:
        # 先尝试转换为浮点数
        num = float(str(value).strip())
        # 如果是整数（如5.0），转换为整数再转字符串
        if num.is_integer():
            return str(int(num))
        else:
            return str(num)
    except (ValueError, TypeError):
        # 如果不是数值，返回原字符串（去除空格）
        return str(value).strip()

def normalize_string(value):
    """字符串标准化"""
    if pd.isna(value):
        return ''
    return str(value).strip()

def read_and_preprocess(file_path, key_columns, sheet_name=0):
    """读取并预处理数据 - 简化版"""
    print(f"📖 读取文件: {os.path.basename(file_path)}")
    
    # 读取Excel文件
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        dtype=str,
        keep_default_na=False,
        engine='openpyxl'
    )
    
    # 清洗列名
    df.columns = df.columns.str.strip().str.replace('\n', '')
    
    # 检查缺失列
    missing_keys = [col for col in key_columns if col not in df.columns]
    if missing_keys:
        raise ValueError(f"缺失主键列: {', '.join(missing_keys)}")
    
    # 预处理所有列
    for col in df.columns:
        df[col] = df[col].apply(normalize_string)
        df[col] = df[col].apply(normalize_number)
    
    # 生成组合主键
    key_parts = []
    for col in key_columns:
        if col in df.columns:
            # 标准化空值处理
            series = df[col].fillna('').astype(str).str.strip()
            series = series.replace(['', 'nan', 'None', 'NaT'], '<空值>')
            key_parts.append(series)
    
    if key_parts:
        df['_composite_key'] = key_parts[0]
        for part in key_parts[1:]:
            df['_composite_key'] = df['_composite_key'] + '_' + part
    else:
        df['_composite_key'] = ''
    
    return df

def compare_datasets(source_df, target_df, original_columns):
    """数据比对函数 - 只核对源文件和目标文件共有的字段"""
    
    # # 处理标准格式（跳过前两行）
    # if len(target_df) > 0 and str(target_df.iloc[0, 0]).strip() == '元数据标准名称':
    #     print('==========标准模板处理===========')
    #     target_df = target_df.iloc[2:].reset_index(drop=True)
    
    # 获取主键映射
    source_key_map = dict(zip(source_df['_composite_key'], source_df.index))
    target_key_map = dict(zip(target_df['_composite_key'], target_df.index))
    
    source_keys = set(source_key_map.keys())
    target_keys = set(target_key_map.keys())
    common_keys = source_keys & target_keys
    
    # 获取源文件和目标文件共有的字段
    source_columns = set(source_df.columns) - {'_composite_key'}
    target_columns = set(target_df.columns) - {'_composite_key'}
    common_columns = source_columns & target_columns
    
    # 按照原始列顺序过滤共有字段
    columns_to_compare = [col for col in original_columns if col in common_columns]
    
    results = []
    
    # 处理仅存在于源数据的主键
    for key in source_keys - target_keys:
        idx = source_key_map[key]
        row_data = {'主键状态': '仅存在于源文件', '组合主键': key}
        
        # 按照原始列顺序添加数据，但只保留共有的字段
        for col in original_columns:
            if col in source_df.columns and col in common_columns:
                row_data[f"源_{col}"] = source_df.at[idx, col]
                row_data[f"目标_{col}"] = ""
        results.append(row_data)
    
    # 处理仅存在于目标数据的主键
    for key in target_keys - source_keys:
        idx = target_key_map[key]
        row_data = {'主键状态': '仅存在于目标文件', '组合主键': key}
        
        # 按照原始列顺序添加数据，但只保留共有的字段
        for col in original_columns:
            if col in target_df.columns and col in common_columns:
                row_data[f"源_{col}"] = ""
                row_data[f"目标_{col}"] = target_df.at[idx, col]
        results.append(row_data)
    
    # 处理共同主键的数据差异
    for key in common_keys:
        src_idx = source_key_map[key]
        tgt_idx = target_key_map[key]
        
        row_data = {'主键状态': '数据一致', '组合主键': key}
        diff_count = 0
        diff_details = {}
        
        # 只比较源文件和目标文件共有的字段
        for col in columns_to_compare:
            src_val = source_df.at[src_idx, col] if col in source_df.columns else None
            tgt_val = target_df.at[tgt_idx, col] if col in target_df.columns else None
            
            # 标准化空值比较
            src_empty = pd.isna(src_val) or str(src_val).strip() in ['', '<空值>']
            tgt_empty = pd.isna(tgt_val) or str(tgt_val).strip() in ['', '<空值>']
            
            row_data[f"源_{col}"] = src_val if src_val is not None else ""
            row_data[f"目标_{col}"] = tgt_val if tgt_val is not None else ""
            
            # 检查差异（两个都非空且不相等）
            if not (src_empty and tgt_empty) and str(src_val) != str(tgt_val):
                diff_count += 1
                diff_details[col] = {'源值': src_val, '目标值': tgt_val}
        
        if diff_count > 0:
            row_data['主键状态'] = f"发现{diff_count}处差异"
            row_data['差异详情'] = str(diff_details)
            row_data['差异列名'] = ",".join(diff_details.keys())
        else:
            row_data['差异详情'] = None
            row_data['差异列名'] = None
        
        results.append(row_data)
    
    return pd.DataFrame(results)

def generate_report(result_df, output_path, original_columns):
    """生成报告 - 保持列顺序，并对差异列进行标红"""
    
    # 构建列顺序：固定列 + 按照原始顺序的源列和目标列
    column_order = ['主键状态', '组合主键']
    
    # 按照原始列顺序添加源列和目标列
    for col in original_columns:
        column_order.append(f"源_{col}")
        column_order.append(f"目标_{col}")
    
    # 添加差异详情列
    column_order.extend(['差异详情', '差异列名'])
    
    # 重新排序（只保留存在的列）
    existing_columns = [col for col in column_order if col in result_df.columns]
    result_df = result_df[existing_columns]
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='核对结果')
        
        # workbook = writer.book
        # worksheet = writer.sheets['核对结果']
        
        # # 定义颜色
        # yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # # 批量标记颜色
        # for row_idx in range(2, len(result_df) + 2):
        #     status = worksheet.cell(row=row_idx, column=1).value
            
        #     if status in ['仅存在于源文件', '仅存在于目标文件']:
        #         # 黄色标记整行
        #         for col_idx in range(1, len(result_df.columns) + 1):
        #             worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
            
        #     elif status and '差异' in str(status):
        #         # 红色标记状态列
        #         worksheet.cell(row=row_idx, column=1).fill = red_fill  # 主键状态
                
        #         # 获取该行的差异列名
        #         diff_cols_cell = worksheet.cell(row=row_idx, column=result_df.columns.get_loc('差异列名') + 1)
        #         if diff_cols_cell.value:
        #             diff_cols = diff_cols_cell.value.split(',')
                    
        #             # 对每个差异列进行标红
        #             for col_name in diff_cols:
        #                 # 查找源列和目标列的索引位置
        #                 source_col_name = f"源_{col_name}"
        #                 target_col_name = f"目标_{col_name}"
                        
        #                 if source_col_name in result_df.columns:
        #                     source_col_idx = result_df.columns.get_loc(source_col_name) + 1
        #                     worksheet.cell(row=row_idx, column=source_col_idx).fill = red_fill
                        
        #                 if target_col_name in result_df.columns:
        #                     target_col_idx = result_df.columns.get_loc(target_col_name) + 1
        #                     worksheet.cell(row=row_idx, column=target_col_idx).fill = red_fill
                
        #         # 红色标记差异详情列
        #         if '差异详情' in result_df.columns:
        #             diff_col_idx = result_df.columns.get_loc('差异详情') + 1
        #             worksheet.cell(row=row_idx, column=diff_col_idx).fill = red_fill
        
        # # 设置列宽
        # for column_cells in worksheet.columns:
        #     max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        #     adjusted_width = min(max_length + 2, 50)
        #     worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        
        # worksheet.freeze_panes = 'A2'

def batch_compare_simple(source_path, target_paths, key_columns, output_dir):
    """简化的批量比对主程序"""
    
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        print(f"\n{'=' * 50} 开始处理 {'=' * 50}")
        start_total = datetime.now()
        
        # 读取源文件并获取原始列顺序
        print(f"[{datetime.now().strftime('%H:%M:%S')}] 读取源文件: {os.path.basename(source_path)}")
        source_df = read_and_preprocess(source_path, key_columns)
        original_columns = [col for col in source_df.columns if col not in ['_composite_key']]
        print(f"源文件读取完成，记录数: {len(source_df):,}")
        
        # 处理每个目标文件
        for i, target_path in enumerate(target_paths, 1):
            start_time = datetime.now()
            target_name = os.path.basename(target_path)
            print(f"\n[{start_time.strftime('%H:%M:%S')}] 处理: {target_name}")
            
            try:
                # 读取目标文件
                target_df = read_and_preprocess(target_path, key_columns)
                print(f"目标文件读取完成，记录数: {len(target_df):,}")
                
                # 比对数据
                result_df = compare_datasets(source_df, target_df, original_columns)
                
                # 生成报告
                report_name = f"比对报告_{os.path.splitext(os.path.basename(source_path))[0]}_vs_{os.path.splitext(target_name)[0]}.xlsx"
                output_path = os.path.join(output_dir, report_name)
                generate_report(result_df, output_path, original_columns)
                
                # 清理目标文件数据缓存
                del target_df
                del result_df
                import gc
                gc.collect()
                print("✅ 目标文件数据缓存已清理")
                
                # 统计信息
                duration = (datetime.now() - start_time).total_seconds()
                
                stats = f"""
✅ 完成比对 ({duration:.2f}秒)
📊 差异统计：
   • 仅存在于源文件: {len(result_df[result_df['主键状态'] == '仅存在于源文件']):,}
   • 仅存在于目标文件: {len(result_df[result_df['主键状态'] == '仅存在于目标文件']):,}
   • 存在数据差异: {len(result_df[result_df['主键状态'].str.contains('差异', na=False)]):,}
   • 总计记录: {len(result_df):,}
💾 报告已保存至: {output_path}
"""
                print(stats)
                
            except Exception as e:
                print(f"❌ 处理失败: {str(e)}")
        
        # 清理源文件数据缓存
        del source_df
        import gc
        gc.collect()
        print("✅ 源文件数据缓存已清理")
        
        total_time = (datetime.now() - start_total).total_seconds()
        print(f"\n🎉 全部处理完成！总耗时: {total_time:.2f}秒")
        
    except Exception as e:
        print(f"\n❌ 严重错误: {str(e)}")
    finally:
        # 最终清理
        import gc
        gc.collect()
        print(f"{'=' * 50} 处理完成 {'=' * 50}")


batch_compare_simple(
    source_path=r'C:\Users\zhangbon\Desktop\MDM物料基础数据-MARA-MARC(1).xlsx',
    target_paths=[
        r'C:\Users\zhangbon\Desktop\SRM系统一致性核对模版-物料20251110.xlsx'
    ],
    key_columns=['工厂','物料编码'],
    output_dir=r'C:\Users\zhangbon\Desktop\比对报告'
)